import io
import json
import zipfile
from datetime import UTC, datetime
from uuid import uuid4

import shapefile
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models.base import Base
from app.models.collection import DataForm, DataFormVersion, Project, Submission, Survey
from app.models.identity import Organization, User
from app.models.operations import MediaEvidence
from app.services.data_export import DataExportService

SQUARE = {
    "type": "Polygon",
    "coordinates": [[[0.0, 0.0], [0.0, 1.0], [1.0, 1.0], [1.0, 0.0], [0.0, 0.0]]],
}


async def _seed():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    session = session_factory()

    org_id, user_id, project_id, survey_id, form_id, version_id = (uuid4() for _ in range(6))
    now = datetime.now(UTC)
    session.add_all(
        [
            Organization(id=org_id, name="Export Org", slug="export-org"),
            User(id=user_id, email="m@example.org", full_name="Manager", password_hash="x"),
            Project(id=project_id, organization_id=org_id, name="P", slug="p", status="active"),
            Survey(id=survey_id, organization_id=org_id, project_id=project_id, created_by_user_id=user_id,
                   owner_user_id=user_id, title="S", code="S1", survey_type="registration", status="active"),
            DataForm(id=form_id, organization_id=org_id, project_id=project_id, survey_id=survey_id,
                     created_by_user_id=user_id, name="Farm Survey", slug="farm-survey", status="published", current_version=1),
            DataFormVersion(id=version_id, organization_id=org_id, form_id=form_id, version=1, offline_compatible=True,
                            published_at=now, schema_json={"sections": [{"id": "s1", "fields": [
                                {"id": "q_name", "variable_name": "name", "type": "text", "label": "Full name"},
                                {"id": "q_boundary", "variable_name": "boundary", "type": "polygon", "label": "Farm boundary"},
                                {"id": "q_loc", "variable_name": "location", "type": "geolocation", "label": "Location"},
                            ]}]}),
        ]
    )

    def _submission(client_id, lat, lng, responses):
        return Submission(
            organization_id=org_id, project_id=project_id, survey_id=survey_id, form_id=form_id,
            form_version_id=version_id, client_submission_id=client_id, status="approved",
            payload_json={"_mobile_responses": responses}, device_id="dev-1",
            captured_at=now, submitted_at=now, sync_received_at=now, latitude=lat, longitude=lng,
            location_captured_at=now,
        )

    poly_sub = _submission("sub-polygon-001", 0.5, 0.5, [
        {"questionId": "q_name", "variableName": "name", "value": "Alice"},
        {"questionId": "q_boundary", "variableName": "boundary", "value": SQUARE},
    ])
    point_sub = _submission("sub-point-002", 1.0, 2.0, [
        {"questionId": "q_name", "variableName": "name", "value": "Bob"},
        {"questionId": "q_loc", "variableName": "location", "value": {"latitude": 1.0, "longitude": 2.0}},
    ])
    session.add_all([poly_sub, point_sub])
    await session.flush()
    session.add(MediaEvidence(
        organization_id=org_id, submission_id=poly_sub.id, form_id=form_id, uploaded_by_user_id=user_id,
        media_type="photo", file_name="farm.jpg", storage_url="https://files/farm.jpg", mime_type="image/jpeg", size_bytes=10,
    ))
    await session.commit()
    return session, org_id, form_id


async def test_capabilities_are_data_aware() -> None:
    session, org_id, form_id = await _seed()
    caps = await DataExportService(session).capabilities(org_id, form_id)
    assert caps is not None
    assert caps["record_count"] == 2
    assert caps["has_polygons"] is True
    assert caps["has_points"] is True
    assert caps["has_media"] is True
    by_id = {fmt["id"]: fmt for fmt in caps["formats"]}
    assert by_id["csv"]["available"] and by_id["xlsx"]["available"] and by_id["json"]["available"]
    assert by_id["geojson"]["available"] and by_id["kml"]["available"]
    assert by_id["shapefile"]["available"] and by_id["gpx"]["available"]


async def test_capabilities_hide_spatial_when_no_location() -> None:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    session = async_sessionmaker(engine, expire_on_commit=False)()
    org_id, user_id, form_id, version_id = (uuid4() for _ in range(4))
    now = datetime.now(UTC)
    session.add_all([
        Organization(id=org_id, name="O", slug="o2"),
        User(id=user_id, email="u@example.org", full_name="U", password_hash="x"),
        DataForm(id=form_id, organization_id=org_id, created_by_user_id=user_id, name="Plain", slug="plain", status="published", current_version=1),
        DataFormVersion(id=version_id, organization_id=org_id, form_id=form_id, version=1, offline_compatible=True, published_at=now,
                        schema_json={"sections": [{"id": "s", "fields": [{"id": "q", "variable_name": "note", "type": "text", "label": "Note"}]}]}),
        Submission(organization_id=org_id, form_id=form_id, form_version_id=version_id, client_submission_id="plain-001",
                   status="approved", payload_json={"_mobile_responses": [{"variableName": "note", "value": "hello"}]},
                   device_id="d", captured_at=now, submitted_at=now, sync_received_at=now, latitude=0, longitude=0, location_captured_at=now),
    ])
    await session.commit()
    caps = await DataExportService(session).capabilities(org_id, form_id)
    by_id = {fmt["id"]: fmt for fmt in caps["formats"]}
    assert by_id["csv"]["available"] is True
    assert by_id["geojson"]["available"] is False
    assert by_id["shapefile"]["available"] is False
    assert by_id["gpx"]["available"] is False


async def test_export_all_formats_produce_valid_output() -> None:
    session, org_id, form_id = await _seed()
    service = DataExportService(session)

    csv_art = await service.export(org_id, form_id, "csv")
    text = csv_art.content.decode("utf-8-sig")
    assert "Full name" in text and "Alice" in text and "geometry_wkt" in text
    assert csv_art.filename.endswith(".csv")

    json_art = await service.export(org_id, form_id, "json")
    parsed = json.loads(json_art.content)
    assert len(parsed["records"]) == 2
    assert any(r["geometry"] and r["geometry"]["type"] == "Polygon" for r in parsed["records"])
    assert any(r["media"] for r in parsed["records"])

    geojson_art = await service.export(org_id, form_id, "geojson")
    fc = json.loads(geojson_art.content)
    assert fc["type"] == "FeatureCollection" and len(fc["features"]) == 2
    assert any(f["geometry"] and f["geometry"]["type"] == "Polygon" for f in fc["features"])
    assert all("Full name" in f["properties"] for f in fc["features"])

    kml = (await service.export(org_id, form_id, "kml")).content.decode("utf-8")
    assert "<Placemark" in kml and "<Polygon>" in kml and "farm.jpg" in kml

    gpx = (await service.export(org_id, form_id, "gpx")).content.decode("utf-8")
    assert gpx.count("<wpt") == 2  # polygon centroid + explicit point

    xlsx_art = await service.export(org_id, form_id, "xlsx")
    workbook = load_workbook(io.BytesIO(xlsx_art.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0][0] == "submission_id" and len(rows) == 3

    shp_art = await service.export(org_id, form_id, "shapefile")
    with zipfile.ZipFile(io.BytesIO(shp_art.content)) as archive:
        names = set(archive.namelist())
        assert {"points.shp", "polygons.shp", "attributes.csv"} <= names
        reader = shapefile.Reader(
            shp=io.BytesIO(archive.read("polygons.shp")),
            shx=io.BytesIO(archive.read("polygons.shx")),
            dbf=io.BytesIO(archive.read("polygons.dbf")),
        )
        assert len(reader.shapes()) == 1
